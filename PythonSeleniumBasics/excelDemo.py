from pathlib import Path
import openpyxl

project_directory = Path(__file__).parent.absolute()
excelDemoFilePath = project_directory / "PythonDemo.xlsx"
book = openpyxl.load_workbook(excelDemoFilePath)
sheet = book.active

Dict = {}

cell = sheet.cell(row=1,column=2)
print(cell.value)

sheet.cell(row=2,column=2).value = "Ravindu"
print(sheet.cell(row=2,column=2).value)
sheet.cell(row=2,column=2).value = "Testfirstname1"

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for rowNumber in range(1,sheet.max_row+1):
    if sheet.cell(row=rowNumber,column=1).value == "Testname1":
        for colNumber in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=colNumber).value] = sheet.cell(row=rowNumber,column=colNumber).value

print(Dict)