from pathlib import Path
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

file_path = Path("C:/Users/Dell/Downloads/download.xlsx")

driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()

# Wait until file is downloaded
for _ in range(15):
    if file_path.exists():
        break
    time.sleep(1)

# Edit the Excel file
book = openpyxl.load_workbook(file_path)
sheet = book.active
Dict = {}
fruit_name = "Apple"
updatingValue = "900"

for columnNo1 in range(1, sheet.max_column + 1):
    if sheet.cell(row=1, column=columnNo1).value == "price":
        Dict["desiredColNo"] = columnNo1
        break

for rowNo2 in range(2, sheet.max_row + 1):
    for columnNo2 in range(1, sheet.max_column + 1):
        if sheet.cell(row=rowNo2, column=columnNo2).value == fruit_name:
            Dict["desiredRowNo"] = rowNo2
            break
    if "desiredRowNo" in Dict:
        break

sheet.cell(row=Dict["desiredRowNo"], column=Dict["desiredColNo"]).value = updatingValue
book.save(file_path)

# Upload the updated Excel file
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(str(file_path))

# Wait for toast message
wait = WebDriverWait(driver, 15)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body")
toast = wait.until(EC.visibility_of_element_located(toast_locator))
toastText = toast.text
assert toastText == "Updated Excel data Successfully."
print("Toast text is: ",toastText)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
assert actual_price == updatingValue
print("Actual price is: ",actual_price)